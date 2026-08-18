"""Unit tests for the Temforce 8-column XLSX writer."""

from datetime import date
from decimal import Decimal

import openpyxl

from stencil.output.mapper import (
    TEMFORCE_OUTPUT_COLUMNS,
    get_line_item_value,
)
from stencil.output.xlsx_writer import (
    allocate_invoice_tax,
    compute_line_tax,
    delivered_tax_sum,
    resolve_tax_rate,
    write_xlsx,
)
from stencil.validation.schema import (
    CanonicalInvoice,
    ExtractionMetadata,
    ExtractionPath,
    InvoiceHeader,
    LineItem,
    OutputType,
)


def _header(**kw) -> InvoiceHeader:
    defaults = dict(
        supplier_name="TestSupplier",
        invoice_number="INV-001",
        invoice_date=date(2025, 12, 2),
        due_date=date(2026, 1, 1),
        account_number="A837737",
        currency="EUR",
    )
    defaults.update(kw)
    return InvoiceHeader(**defaults)


def _meta() -> ExtractionMetadata:
    return ExtractionMetadata(extraction_path=ExtractionPath.AI)


def _li(amount: str, charge_type: str = "recurring", **kw) -> LineItem:
    return LineItem(
        line_number=kw.pop("line_number", 1),
        service_id=kw.pop("service_id", "S00364341"),
        billing_reference=kw.pop("billing_reference", "C35332-338"),
        description=kw.pop("description", "Test service"),
        charge_type=charge_type,
        amount=Decimal(amount),
        **kw,
    )


def _invoice(line_items=None, **kw) -> CanonicalInvoice:
    items = line_items or [_li("850.00"), _li("600.00", line_number=2)]
    defaults = dict(
        intake_id="test-001",
        output_type=OutputType.STANDARD,
        header=_header(),
        line_items=items,
        subtotal=Decimal("1450.00"),
        tax=Decimal("333.50"),
        total_due=Decimal("1783.50"),
        metadata=_meta(),
    )
    defaults.update(kw)
    return CanonicalInvoice(**defaults)


def _with_calculated_tax(inv: CanonicalInvoice, source: str = "auto") -> CanonicalInvoice:
    inv.fields["_tax_output_mode"] = "calculate"
    inv.fields["_tax_rate_source"] = source
    return inv


# ── Column definitions ───────────────────────────────────────


class TestTemforceColumns:
    def test_exactly_8_columns(self):
        assert len(TEMFORCE_OUTPUT_COLUMNS) == 8

    def test_column_names(self):
        names = [c.xlsx_header for c in TEMFORCE_OUTPUT_COLUMNS]
        assert names == [
            "EXT_SERVICEID", "EXT_BILLINGREFERENCE", "EXT_DATE", "formula",
            "EXT_AMOUNT", "EXT_ACCOUNT", "EXT_INVOICENUMBER", "EXT_TAX",
        ]


# ── Fallback logic ───────────────────────────────────────────


class TestFallbackPath:
    def test_uses_billing_reference_when_present(self):
        li = {"billing_reference": "C35332-338", "service_id": "S00364341"}
        value = get_line_item_value(
            "line_item.billing_reference", li, {}, {},
            fallback_path="line_item.service_id",
        )
        assert value == "C35332-338"

    def test_falls_back_to_service_id(self):
        li = {"billing_reference": None, "service_id": "S00364341"}
        value = get_line_item_value(
            "line_item.billing_reference", li, {}, {},
            fallback_path="line_item.service_id",
        )
        assert value == "S00364341"

    def test_no_fallback_when_not_set(self):
        li = {"billing_reference": None, "service_id": "S00364341"}
        value = get_line_item_value(
            "line_item.billing_reference", li, {}, {},
        )
        assert value is None

    def test_header_values_resolve(self):
        header = {"invoice_date": "2025-12-02", "account_number": "A837737"}
        value = get_line_item_value("header.invoice_date", {}, header, {})
        assert value == "2025-12-02"

    def test_header_fallback_resolves(self):
        header = {"invoice_date": "2025-12-02", "due_date": None}
        value = get_line_item_value(
            "field.due_date",
            {},
            header,
            {},
            fallback_path="field.invoice_date",
        )
        assert value == "2025-12-02"

    def test_computed_returns_none(self):
        value = get_line_item_value("computed.line_tax", {}, {}, {})
        assert value is None


# ── Tax rate resolution ──────────────────────────────────────


class TestTaxRate:
    def test_uses_extracted_tax_rate(self):
        inv = _invoice(tax_rate=Decimal("0.23"))
        rate = resolve_tax_rate(inv)
        assert rate == Decimal("0.23")

    def test_normalizes_extracted_whole_percent_tax_rate(self):
        inv = _invoice(tax_rate=Decimal("19"))
        rate = resolve_tax_rate(inv)
        assert rate == Decimal("0.19")

    def test_derives_from_totals(self):
        inv = _invoice()  # subtotal=1450, tax=333.50
        rate = resolve_tax_rate(inv)
        assert rate == Decimal("333.50") / Decimal("1450.00")

    def test_returns_none_when_nothing_available(self):
        inv = _invoice(subtotal=None, tax=None)
        rate = resolve_tax_rate(inv)
        assert rate is None

    def test_derives_from_consistent_per_line_tax(self):
        items = [
            _li("3471.19", tax_amount=Decimal("867.80")),
            _li("3052.08", tax_amount=Decimal("763.02"), line_number=2),
            _li("5460.00", tax_amount=Decimal("1365.00"), line_number=3),
        ]
        inv = _invoice(line_items=items, subtotal=None, tax=None, total_due=None)
        rate = resolve_tax_rate(inv)
        assert rate == Decimal("0.25")

    def test_inconsistent_per_line_tax_does_not_become_rate(self):
        items = [
            _li("100.00", tax_amount=Decimal("10.00")),
            _li("100.00", tax_amount=Decimal("23.00"), line_number=2),
        ]
        inv = _invoice(line_items=items, subtotal=None, tax=None, total_due=None)
        rate = resolve_tax_rate(inv)
        assert rate is None

    def test_extracted_rate_takes_precedence(self):
        inv = _invoice(tax_rate=Decimal("0.19"))  # 19% overrides derived 23%
        rate = resolve_tax_rate(inv)
        assert rate == Decimal("0.19")


# ── Row filtering ────────────────────────────────────────────


class TestRowFiltering:
    def test_tax_rows_excluded(self, tmp_path):
        items = [
            _li("850.00", "recurring", line_number=1),
            _li("195.50", "tax", line_number=2, service_id=None, billing_reference=None,
                description="VAT 23%"),
        ]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("0.23")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row (tax row excluded)
        assert float(ws.cell(row=2, column=5).value) == 850.00  # EXT_AMOUNT

    def test_standalone_fee_and_surcharge_rows_excluded(self, tmp_path):
        items = [
            _li("850.00", "recurring", line_number=1),
            _li(
                "50.00", "fee", line_number=2, service_id=None,
                billing_reference=None, description="Admin fee",
            ),
            _li(
                "25.00", "surcharge", line_number=3, service_id=None,
                billing_reference=None, description="Fuel surcharge",
            ),
        ]
        inv = _invoice(line_items=items)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 service row

    def test_service_backed_fee_row_is_delivered(self, tmp_path):
        items = [
            _li(
                "1000.22", "fee", service_id="P850305305",
                billing_reference="Administrative Charge", tax_amount=Decimal("0"),
            ),
        ]
        inv = _invoice(line_items=items, subtotal=Decimal("1000.22"), tax=Decimal("0"))
        inv.fields["_tax_output_mode"] = "extract_exact"
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "P850305305"
        assert ws.cell(row=2, column=2).value == "Administrative Charge"
        assert ws.cell(row=2, column=5).value == 1000.22
        assert ws.cell(row=2, column=8).value == 0

    def test_service_backed_surcharge_row_is_delivered(self, tmp_path):
        items = [
            _li(
                "750.00", "surcharge", service_id="SGOSGO197798",
                billing_reference="SGOSGO197798", tax_amount=Decimal("0"),
                description="TDM Access Surcharge",
            ),
        ]
        inv = _invoice(line_items=items, subtotal=Decimal("750"), tax=Decimal("0"))
        inv.fields["_tax_output_mode"] = "extract_exact"
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "SGOSGO197798"
        assert ws.cell(row=2, column=5).value == 750
        assert ws.cell(row=2, column=8).value == 0


# ── XLSX output format ───────────────────────────────────────


class TestXlsxOutput:
    def test_writes_8_columns(self, tmp_path):
        inv = _with_calculated_tax(_invoice(tax_rate=Decimal("0.23")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 9)]
        assert headers == [
            "EXT_SERVICEID", "EXT_BILLINGREFERENCE", "EXT_DATE", "formula",
            "EXT_AMOUNT", "EXT_ACCOUNT", "EXT_INVOICENUMBER", "EXT_TAX",
        ]
        assert ws.cell(row=1, column=9).value is None  # no 9th column

    def test_per_line_tax_computed(self, tmp_path):
        items = [_li("850.00")]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("0.23")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        tax_value = ws.cell(row=2, column=8).value  # EXT_TAX
        assert tax_value == 195.50  # 850 * 0.23

    def test_ext_amount_is_numeric_not_string(self, tmp_path):
        from stencil.output.xlsx_writer import build_output_rows

        inv = _with_calculated_tax(_invoice(line_items=[_li("850.00")], tax_rate=Decimal("0.23")))
        rows = build_output_rows(inv)
        amount_col = next(
            i for i, col in enumerate(TEMFORCE_OUTPUT_COLUMNS)
            if col.xlsx_header == "EXT_AMOUNT"
        )
        assert isinstance(rows[0][amount_col], float)
        assert rows[0][amount_col] == 850.0

        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)
        wb = openpyxl.load_workbook(out)
        cell = wb.active.cell(row=2, column=amount_col + 1)
        assert isinstance(cell.value, (int, float))
        assert float(cell.value) == 850.0

    def test_service_id_and_billing_reference_separate(self, tmp_path):
        items = [_li("100.00", billing_reference="C99-123", service_id="S001")]
        inv = _invoice(line_items=items)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "S001"  # EXT_SERVICEID
        assert ws.cell(row=2, column=2).value == "C99-123"  # EXT_BILLINGREFERENCE

    def test_falls_back_to_service_id(self, tmp_path):
        items = [_li("100.00", billing_reference=None, service_id="S001")]
        inv = _invoice(line_items=items)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "S001"
        assert ws.cell(row=2, column=2).value == "S001"

    def test_dates_formatted_mm_dd_yyyy(self, tmp_path):
        inv = _with_calculated_tax(_invoice(tax_rate=Decimal("0.23")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "12/02/2025"  # EXT_DATE
        assert ws.cell(row=2, column=4).value == "01/01/2026"  # formula (due date)

    def test_formula_falls_back_to_invoice_date_when_due_date_missing(self, tmp_path):
        inv = _with_calculated_tax(_invoice(header=_header(due_date=None), tax_rate=Decimal("0.23")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "12/02/2025"
        assert ws.cell(row=2, column=4).value == "12/02/2025"

    def test_header_fields_repeated(self, tmp_path):
        items = [_li("100.00", line_number=1), _li("200.00", line_number=2)]
        inv = _invoice(line_items=items)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Both rows should have the same header values
        for row in [2, 3]:
            assert ws.cell(row=row, column=6).value == "A837737"  # EXT_ACCOUNT
            assert ws.cell(row=row, column=7).value == "INV-001"  # EXT_INVOICENUMBER

    def test_no_tax_rate_leaves_empty(self, tmp_path):
        items = [_li("1000.00")]
        inv = _invoice(line_items=items, subtotal=None, tax=None, tax_rate=None)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=8).value is None


class TestLineTax:
    def test_uses_per_line_tax_when_present(self, tmp_path):
        items = [_li("3071.00", tax_amount=Decimal("2020.81"))]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("0.09")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=8).value == 2020.81  # extracted, not 3071*0.09

    def test_per_line_tax_takes_precedence_over_rate(self):
        li = _li("1000.00", tax_amount=Decimal("999.99"))
        assert compute_line_tax(li, Decimal("0.23")) == 999.99

    def test_calculates_from_rate_when_line_tax_missing(self, tmp_path):
        items = [_li("1000.00", tax_amount=None)]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("0.10")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=8).value == 100.0  # 1000 * 0.10

    def test_calculates_from_whole_percent_rate_when_line_tax_missing(self, tmp_path):
        items = [_li("72.00", tax_amount=None)]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("19")))
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=8).value == 13.68  # 72 * 0.19

    def test_explicit_tax_policy_suppresses_rate_fallback(self, tmp_path):
        items = [
            _li("1000.00", tax_amount=None),
            _li("2000.00", tax_amount=Decimal("123.45"), line_number=2),
        ]
        inv = _with_calculated_tax(_invoice(line_items=items, tax_rate=Decimal("0.10")))
        inv.fields["_tax_output_mode"] = "extract_exact"
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=8).value is None

    def test_default_auto_allocates_invoice_level_tax(self, tmp_path):
        # Default mode is now "auto": with no per-line tax but an invoice-level
        # tax total, the tax is allocated across lines (sums exactly to the total).
        items = [_li("1000.00", tax_amount=None)]
        inv = _invoice(line_items=items, subtotal=Decimal("1000.00"), tax=Decimal("120.00"), tax_rate=None)
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        assert wb.active.cell(row=2, column=8).value == 120.0
        assert inv.fields["_tax_method"] == "allocated"

    def test_calculate_from_invoice_tax_divided_by_subtotal(self, tmp_path):
        items = [_li("1000.00", tax_amount=None)]
        inv = _with_calculated_tax(
            _invoice(line_items=items, subtotal=Decimal("1000.00"), tax=Decimal("120.00"), tax_rate=None),
            source="invoice_tax_divided_by_subtotal",
        )
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        assert wb.active.cell(row=2, column=8).value == 120.0

    def test_none_tax_output_mode_blanks_even_raw_tax(self, tmp_path):
        items = [_li("1000.00", tax_amount=Decimal("120.00"))]
        inv = _invoice(line_items=items)
        inv.fields["_tax_output_mode"] = "none"
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        assert wb.active.cell(row=2, column=8).value is None

    def test_bell_mts_account_tax_not_allocated_to_subscribers(self, tmp_path):
        items = [
            _li("0", tax_amount=Decimal("56.59"), service_id="Account Level", billing_reference=None),
            _li("34.22", tax_amount=None, line_number=2, service_id="204 889-8280"),
            _li("40.42", tax_amount=None, line_number=3, service_id="204 925-6602"),
        ]
        inv = _invoice(line_items=items, subtotal=Decimal("471.62"), tax=Decimal("56.59"), tax_rate=None)
        inv.fields["_tax_output_mode"] = "extract_exact"
        out = tmp_path / "test.xlsx"
        write_xlsx(inv, out)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        tax_values = [ws.cell(row=i, column=8).value for i in range(2, 5)]
        assert tax_values == [56.59, None, None]
        assert sum(Decimal(str(value or 0)) for value in tax_values) == Decimal("56.59")


class TestAllocateInvoiceTax:
    def _items(self, *amounts):
        return [_li(a, tax_amount=None, line_number=i + 1) for i, a in enumerate(amounts)]

    def test_single_line_gets_full_tax(self):
        inv = _invoice(line_items=self._items("1000.00"), subtotal=Decimal("1000"), tax=Decimal("120.00"))
        assert allocate_invoice_tax(inv) == [Decimal("120.00")]

    def test_sums_exactly_with_penny_residual(self):
        # 10.00 tax over three equal lines cannot divide evenly to cents.
        inv = _invoice(
            line_items=self._items("100", "100", "100"),
            subtotal=Decimal("300"),
            tax=Decimal("10.00"),
        )
        alloc = allocate_invoice_tax(inv)
        assert sum(alloc) == Decimal("10.00")
        assert all(a.quantize(Decimal("0.01")) == a for a in alloc)

    def test_proportional_to_amount(self):
        inv = _invoice(
            line_items=self._items("750", "250"),
            subtotal=Decimal("1000"),
            tax=Decimal("100.00"),
        )
        assert allocate_invoice_tax(inv) == [Decimal("75.00"), Decimal("25.00")]

    def test_zero_tax_allocates_nothing(self):
        inv = _invoice(line_items=self._items("100", "200"), subtotal=Decimal("300"), tax=Decimal("0"))
        assert allocate_invoice_tax(inv) == [Decimal("0"), Decimal("0")]

    def test_none_tax_allocates_nothing(self):
        inv = _invoice(line_items=self._items("100"), subtotal=Decimal("100"))
        inv.fields.pop("tax", None)
        assert all(v == Decimal("0") for v in allocate_invoice_tax(inv))

    def test_zero_amount_sum_does_not_crash(self):
        inv = _invoice(line_items=self._items("0", "0"), subtotal=Decimal("0"), tax=Decimal("5.00"))
        assert allocate_invoice_tax(inv) == [Decimal("0"), Decimal("0")]

    def test_credit_line_gets_negative_tax_and_sum_holds(self):
        inv = _invoice(
            line_items=self._items("1000", "-200"),
            subtotal=Decimal("800"),
            tax=Decimal("80.00"),
        )
        alloc = allocate_invoice_tax(inv)
        assert sum(alloc) == Decimal("80.00")
        assert alloc[1] < 0  # the credit line carries negative tax


class TestAutoTaxResolver:
    def _auto(self, **kw):
        inv = _invoice(**kw)
        inv.fields["_tax_output_mode"] = "auto"
        return inv

    def test_auto_uses_printed_per_line_tax(self, tmp_path):
        items = [_li("1000.00", tax_amount=Decimal("90.00")), _li("500.00", tax_amount=None, line_number=2)]
        inv = self._auto(line_items=items, subtotal=Decimal("1500"), tax=Decimal("90.00"))
        out = tmp_path / "t.xlsx"
        write_xlsx(inv, out)
        ws = openpyxl.load_workbook(out).active
        assert [ws.cell(row=i, column=8).value for i in (2, 3)] == [90.0, None]
        assert inv.fields["_tax_method"] == "per_line"

    def test_auto_allocates_invoice_level_tax(self, tmp_path):
        items = [_li("600.00", tax_amount=None), _li("400.00", tax_amount=None, line_number=2)]
        inv = self._auto(line_items=items, subtotal=Decimal("1000"), tax=Decimal("100.00"))
        out = tmp_path / "t.xlsx"
        write_xlsx(inv, out)
        ws = openpyxl.load_workbook(out).active
        assert [ws.cell(row=i, column=8).value for i in (2, 3)] == [60.0, 40.0]
        assert delivered_tax_sum(inv) == Decimal("100.00")
        assert inv.fields["_tax_method"] == "allocated"

    def test_auto_falls_back_to_rate(self, tmp_path):
        # No per-line tax and no invoice-level tax total, but a printed rate.
        items = [_li("1000.00", tax_amount=None)]
        inv = self._auto(line_items=items, subtotal=Decimal("1000"), tax_rate=Decimal("0.10"))
        inv.fields.pop("tax", None)
        out = tmp_path / "t.xlsx"
        write_xlsx(inv, out)
        ws = openpyxl.load_workbook(out).active
        assert ws.cell(row=2, column=8).value == 100.0
        assert inv.fields["_tax_method"] == "rate"

    def test_auto_blank_when_no_tax_evidence(self, tmp_path):
        items = [_li("1000.00", tax_amount=None)]
        inv = self._auto(line_items=items, subtotal=Decimal("1000"), tax_rate=None)
        inv.fields.pop("tax", None)
        out = tmp_path / "t.xlsx"
        write_xlsx(inv, out)
        ws = openpyxl.load_workbook(out).active
        assert ws.cell(row=2, column=8).value is None
        assert inv.fields["_tax_method"] == "none"
