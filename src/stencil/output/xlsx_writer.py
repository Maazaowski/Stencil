"""Generate Temforce-ready XLSX files from canonical invoice data."""

from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from stencil.output.mapper import (
    TEMFORCE_OUTPUT_COLUMNS,
    apply_output_transforms,
    normalize_output_value,
    output_spec_to_columns,
    resolve_output_cell,
)
from stencil.output.spec import OutputSpec
from stencil.validation.schema import CanonicalInvoice, ChargeType, LineItem

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)

ZERO = Decimal("0")
ONE = Decimal("1")
ONE_HUNDRED = Decimal("100")
CENT = Decimal("0.01")
TAX_OUTPUT_EXTRACT_EXACT = "extract_exact"
TAX_OUTPUT_CALCULATE = "calculate"
TAX_OUTPUT_NONE = "none"
TAX_OUTPUT_AUTO = "auto"
TAX_RATE_AUTO = "auto"
TAX_RATE_INVOICE_RATE = "invoice_tax_rate"
TAX_RATE_INVOICE_TOTALS = "invoice_tax_divided_by_subtotal"
TAX_RATE_CONSISTENT_LINES = "consistent_line_tax_rate"

# Tax is always a delivered column rather than a standalone row. Fees and
# surcharges are excluded only when they are invoice-level adjustments without
# a service identifier; a service-backed row is an invoice-backed line item.
_ALWAYS_EXCLUDED_CHARGE_TYPES = frozenset({ChargeType.TAX})
_VALID_TAX_OUTPUT_MODES = frozenset({
    TAX_OUTPUT_EXTRACT_EXACT,
    TAX_OUTPUT_CALCULATE,
    TAX_OUTPUT_NONE,
    TAX_OUTPUT_AUTO,
})
_VALID_TAX_RATE_SOURCES = frozenset({
    TAX_RATE_AUTO,
    TAX_RATE_INVOICE_RATE,
    TAX_RATE_INVOICE_TOTALS,
    TAX_RATE_CONSISTENT_LINES,
})


def _has_service_id(line_item: LineItem) -> bool:
    value = str(line_item.service_id or "").strip()
    return bool(value and any(char.isalnum() for char in value))


def _is_deliverable_adjustment(line_item: LineItem) -> bool:
    if line_item.charge_type == ChargeType.FEE:
        return _has_service_id(line_item)
    if line_item.charge_type == ChargeType.SURCHARGE:
        # Compact extraction may carry the preceding service ID onto a standalone
        # invoice-level surcharge summary. Require explicit per-line tax evidence
        # (including a printed zero) as well as the identifier before delivering
        # it as an invoice-backed charge row.
        return _has_service_id(line_item) and line_item.tax_amount is not None
    return True


def service_line_items(invoice: CanonicalInvoice) -> list[LineItem]:
    """Return rows that appear in the Temforce output.

    Standalone tax rows remain excluded. Standalone fee and surcharge rows are
    also excluded, but either type carrying a real service ID is delivered
    because it represents an invoice-backed service block.
    """
    return [
        line_item
        for line_item in invoice.line_items
        if line_item.charge_type not in _ALWAYS_EXCLUDED_CHARGE_TYPES
        and _is_deliverable_adjustment(line_item)
    ]


def service_items_have_per_line_tax(invoice: CanonicalInvoice) -> bool:
    """True when any output row carries an extracted per-line tax amount."""
    return any(li.tax_amount is not None for li in service_line_items(invoice))


def normalize_tax_rate(tax_rate: Decimal | None) -> Decimal | None:
    """Return tax rates as decimal multipliers; accept extracted whole percents."""
    if tax_rate is None:
        return None
    if tax_rate > ONE:
        return tax_rate / ONE_HUNDRED
    return tax_rate


def tax_output_mode(invoice: CanonicalInvoice) -> str:
    """Profile-controlled delivered EXT_TAX policy."""
    if invoice.fields.get("_tax_fallback_policy") == "explicit_only":
        return TAX_OUTPUT_EXTRACT_EXACT
    raw = str(invoice.fields.get("_tax_output_mode") or TAX_OUTPUT_AUTO).strip()
    return raw if raw in _VALID_TAX_OUTPUT_MODES else TAX_OUTPUT_AUTO


def tax_rate_source(invoice: CanonicalInvoice) -> str:
    raw = str(invoice.fields.get("_tax_rate_source") or TAX_RATE_AUTO).strip()
    return raw if raw in _VALID_TAX_RATE_SOURCES else TAX_RATE_AUTO


def _explicit_invoice_tax_rate(invoice: CanonicalInvoice) -> Decimal | None:
    if invoice.tax_rate is not None:
        return normalize_tax_rate(invoice.tax_rate)
    return None


def _tax_rate_from_invoice_totals(invoice: CanonicalInvoice) -> Decimal | None:
    if invoice.tax is not None and invoice.subtotal is not None and invoice.subtotal > ZERO:
        return invoice.tax / invoice.subtotal
    return None


def _tax_rate_from_consistent_lines(invoice: CanonicalInvoice) -> Decimal | None:
    rates: list[Decimal] = []
    for item in service_line_items(invoice):
        if item.tax_amount is None or item.amount in (None, ZERO):
            continue
        rates.append((item.tax_amount / item.amount).quantize(Decimal("0.0001")))
    if rates and all(rate == rates[0] for rate in rates):
        return rates[0].normalize()
    return None


def resolve_tax_rate(invoice: CanonicalInvoice, source: str = TAX_RATE_AUTO) -> Decimal | None:
    """Resolve a tax rate from the allowed source for calculated EXT_TAX."""
    source = source if source in _VALID_TAX_RATE_SOURCES else TAX_RATE_AUTO
    if source == TAX_RATE_INVOICE_RATE:
        return _explicit_invoice_tax_rate(invoice)
    if source == TAX_RATE_INVOICE_TOTALS:
        return _tax_rate_from_invoice_totals(invoice)
    if source == TAX_RATE_CONSISTENT_LINES:
        return _tax_rate_from_consistent_lines(invoice)

    for resolver in (_explicit_invoice_tax_rate, _tax_rate_from_invoice_totals, _tax_rate_from_consistent_lines):
        rate = resolver(invoice)
        if rate is not None:
            return rate
    return None


def compute_line_tax(line_item: LineItem, tax_rate: Decimal | None) -> float | None:
    """Compute EXT_TAX for one service line from invoice data.

    Uses the per-line tax amount when the invoice states one; otherwise
    calculates amount × invoice tax rate when a rate is available.
    """
    if line_item.tax_amount is not None:
        return round(float(line_item.tax_amount.quantize(CENT, rounding=ROUND_HALF_UP)), 2)
    if tax_rate is not None:
        return float((line_item.amount * tax_rate).quantize(CENT, rounding=ROUND_HALF_UP))
    return None


def output_tax_rate(invoice: CanonicalInvoice) -> Decimal | None:
    if tax_output_mode(invoice) != TAX_OUTPUT_CALCULATE:
        return None
    return resolve_tax_rate(invoice, tax_rate_source(invoice))


def allocate_invoice_tax(invoice: CanonicalInvoice) -> list[Decimal]:
    """Distribute the invoice-level tax total across output lines so the per-line
    taxes sum EXACTLY to ``invoice.tax``.

    Returns a list aligned to ``service_line_items(invoice)`` order (NOT keyed by
    object — ``invoice.line_items`` rebuilds fresh LineItem objects on every
    access, so positional alignment is the only stable join).

    Allocation is proportional to each line's (signed) amount, using cumulative
    rounding: each line's tax is the running rounded cumulative tax minus the
    previous cumulative. This telescopes to exactly ``invoice.tax`` with no penny
    drift, is deterministic (line order), and handles credits (negative amounts
    get negative tax).
    """
    total_tax = invoice.tax
    items = service_line_items(invoice)
    if total_tax is None or total_tax == ZERO or not items:
        return [ZERO for _ in items]
    amount_sum = sum((li.amount for li in items), ZERO)
    if amount_sum == ZERO:
        return [ZERO for _ in items]

    allocations: list[Decimal] = []
    running_amount = ZERO
    prev_cum_tax = ZERO
    for line_item in items:
        running_amount += line_item.amount
        cum_tax = (total_tax * running_amount / amount_sum).quantize(CENT, rounding=ROUND_HALF_UP)
        allocations.append(cum_tax - prev_cum_tax)
        prev_cum_tax = cum_tax
    return allocations


def _auto_tax_plan(invoice: CanonicalInvoice) -> tuple[str, list[Decimal] | None, Decimal | None]:
    """Resolve the auto tax strategy once per delivery.

    Returns (method, allocation_list, rate) where method is one of
    ``per_line`` / ``allocated`` / ``rate`` / ``none``. ``allocation_list`` is
    aligned to ``service_line_items(invoice)`` order.
    """
    if service_items_have_per_line_tax(invoice):
        return ("per_line", None, None)
    total_tax = invoice.tax
    amount_sum = sum((li.amount for li in service_line_items(invoice)), ZERO)
    if total_tax is not None and total_tax != ZERO and amount_sum != ZERO:
        return ("allocated", allocate_invoice_tax(invoice), None)
    rate = resolve_tax_rate(invoice, tax_rate_source(invoice))
    if rate is not None:
        return ("rate", None, rate)
    return ("none", None, None)


def _delivered_line_taxes(invoice: CanonicalInvoice) -> list[float | None]:
    """The final EXT_TAX per output line for every mode, computed once.

    Returns a list aligned to ``service_line_items(invoice)`` order. Also stashes
    the resolved strategy on ``invoice.fields['_tax_method']`` so it can surface
    in the manifest and results UI.
    """
    mode = tax_output_mode(invoice)
    items = service_line_items(invoice)

    if mode == TAX_OUTPUT_AUTO:
        method, allocation, rate = _auto_tax_plan(invoice)
        invoice.fields["_tax_method"] = method
        result: list[float | None] = []
        for idx, line_item in enumerate(items):
            if method == "per_line":
                result.append(compute_line_tax(line_item, None))
            elif method == "allocated":
                value = allocation[idx] if allocation is not None and idx < len(allocation) else ZERO
                result.append(float(value.quantize(CENT, rounding=ROUND_HALF_UP)))
            elif method == "rate":
                result.append(compute_line_tax(line_item, rate))
            else:
                result.append(None)
        return result

    # Legacy explicit modes.
    tax_rate = output_tax_rate(invoice)
    invoice.fields["_tax_method"] = {
        TAX_OUTPUT_EXTRACT_EXACT: "per_line",
        TAX_OUTPUT_CALCULATE: "rate",
        TAX_OUTPUT_NONE: "none",
    }.get(mode, mode)
    return [line_tax_value(li, invoice, tax_rate) for li in items]


def line_tax_value(line_item: LineItem, invoice: CanonicalInvoice, tax_rate: Decimal | None) -> float | None:
    """Per-line EXT_TAX for the explicit modes.

    Auto allocation is cross-line and resolved in ``_delivered_line_taxes`` /
    ``build_output_rows``; auto falls through here to per-line printed tax only
    (this path is not used for auto delivery).
    """
    mode = tax_output_mode(invoice)
    if mode == TAX_OUTPUT_NONE:
        return None
    if mode in (TAX_OUTPUT_EXTRACT_EXACT, TAX_OUTPUT_AUTO):
        return compute_line_tax(line_item, None)
    return compute_line_tax(line_item, tax_rate)


def delivered_tax_sum(invoice: CanonicalInvoice) -> Decimal:
    """Sum taxes exactly as delivered through computed.line_tax."""
    total = ZERO
    for value in _delivered_line_taxes(invoice):
        if value is not None:
            total += Decimal(str(value))
    return total.quantize(CENT, rounding=ROUND_HALF_UP)


# Registry of derived ("computed.<name>") output fields. Each takes the line
# item, the whole invoice, and the resolved invoice tax rate. New customer specs
# can reference these without changing the row builder.
COMPUTED_FIELDS = {
    "line_tax": line_tax_value,
}


def build_output_rows(invoice: CanonicalInvoice, spec: OutputSpec | None = None) -> list[list]:
    """Produce the delivered output rows (one list of column values per service line).

    This is the single source of truth for what the delivered XLSX contains, so
    it can be reused to compare two extractions for an exact, deliverable-level
    match (e.g. AI output vs. model output during candidate validation). The
    column set comes from ``spec`` (the customer's OutputSpec); when omitted it
    falls back to the built-in TemForce 8-column layout.
    """
    columns = output_spec_to_columns(spec) if spec is not None else TEMFORCE_OUTPUT_COLUMNS
    tax_rate = output_tax_rate(invoice)
    items = service_line_items(invoice)
    # Resolve delivered per-line taxes once (handles auto allocation without
    # recomputing the cross-line plan per row). Aligned positionally to ``items``
    # because ``invoice.line_items`` rebuilds fresh objects on each access.
    line_taxes = _delivered_line_taxes(invoice)

    rows: list[list] = []
    for idx, line_item in enumerate(items):
        row: list = []
        for col_def in columns:
            path = col_def.canonical_path
            if path.startswith("computed."):
                computed_name = path.removeprefix("computed.")
                if computed_name == "line_tax":
                    value = line_taxes[idx] if idx < len(line_taxes) else None
                else:
                    fn = COMPUTED_FIELDS.get(computed_name)
                    value = fn(line_item, invoice, tax_rate) if fn else None
            else:
                value = resolve_output_cell(
                    path,
                    line_item,
                    invoice,
                    fallback_path=col_def.fallback_path,
                )
            value = apply_output_transforms(value, col_def.transforms)
            row.append(normalize_output_value(value, path))
        rows.append(row)
    return rows


def write_xlsx(
    invoice: CanonicalInvoice,
    output_path: Path,
    spec: OutputSpec | None = None,
) -> Path:
    """Write a canonical invoice to a deliverable XLSX.

    Columns come from ``spec`` (the customer's OutputSpec); when omitted it falls
    back to the built-in TemForce 8-column layout. Derived columns (e.g. EXT_TAX
    via ``computed.line_tax``) are resolved from the invoice automatically.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    columns = output_spec_to_columns(spec) if spec is not None else TEMFORCE_OUTPUT_COLUMNS

    # Write header row
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def.xlsx_header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width

    # Write data rows from the shared row builder
    rows = build_output_rows(invoice, spec)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (col_def, value) in enumerate(zip(columns, row), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            if col_def.number_format:
                cell.number_format = col_def.number_format

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
