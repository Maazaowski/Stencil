"""Parse a user-supplied XLS/XLSX blueprint into deliverable rows.

During AI profile authoring the user can attach a historical deliverable for a
sample invoice. We treat that workbook as a blueprint: it teaches the output
shape and historical mapping behavior, while the invoice remains the source of
truth for financial values.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

import structlog
from openpyxl import load_workbook

from stencil.output.mapper import TEMFORCE_OUTPUT_COLUMNS, output_spec_to_columns
from stencil.output.spec import OutputSpec

logger = structlog.get_logger()


def _norm_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _cell_text(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return "" if value is None else str(value).strip()


def _decimal_cell(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        cleaned = str(value).strip().replace("$", "").replace(",", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = f"-{cleaned[1:-1]}"
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _workbook_rows(path: Path) -> tuple[str, list[list[object]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            ws = wb.active
            return ws.title, [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - deployment guard
            raise RuntimeError("Legacy .xls blueprint support requires xlrd>=2.0.1") from exc
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        return sheet.name, rows
    raise ValueError(f"Unsupported blueprint workbook suffix: {suffix}")


def _sum_column(rows: list[list[object]], header_row: list[str], names: set[str]) -> str | None:
    indexes = [i for i, header in enumerate(header_row) if header in names]
    total = Decimal("0")
    seen = False
    for row in rows:
        for idx in indexes:
            amount = _decimal_cell(row[idx] if idx < len(row) else None)
            if amount is not None:
                total += amount
                seen = True
    return str(total) if seen else None


def load_blueprint_context(path: str | Path, spec: OutputSpec | None, *, sample_rows: int = 10) -> dict:
    """Read a blueprint workbook and return aligned rows plus prompt context.

    The first non-empty row is treated as the header. Columns are matched to the
    active output spec by header text. If no headers match, columns are taken
    positionally so headerless or odd legacy exports can still provide examples.
    """
    workbook_path = Path(path)
    columns = output_spec_to_columns(spec) if spec is not None else list(TEMFORCE_OUTPUT_COLUMNS)
    spec_headers = [_norm_header(c.xlsx_header) for c in columns]
    sheet_name, all_rows = _workbook_rows(workbook_path)
    warnings: list[str] = []

    header_idx = next(
        (i for i, row in enumerate(all_rows) if any(c is not None and str(c).strip() for c in row)),
        None,
    )
    if header_idx is None:
        return {
            "filename": workbook_path.name,
            "sheet_name": sheet_name,
            "raw_headers": [],
            "aligned_headers": [c.xlsx_header for c in columns],
            "aligned_rows": [],
            "row_count": 0,
            "sample_rows": [],
            "totals": {},
            "warnings": ["Blueprint workbook is empty."],
            "contract_compatible": False,
            "missing_output_headers": [c.xlsx_header for c in columns],
            "unexpected_headers": [],
        }

    raw_header_row = all_rows[header_idx]
    header_row = [_norm_header(c) for c in raw_header_row]
    raw_headers = [_cell_text(c) for c in raw_header_row]

    col_index: list[int | None] = []
    missing_headers: list[str] = []
    for pos, header in enumerate(spec_headers):
        if header and header in header_row:
            col_index.append(header_row.index(header))
        elif not any(h for h in header_row):
            col_index.append(pos if pos < len(header_row) else None)
        else:
            col_index.append(None)
            missing_headers.append(columns[pos].xlsx_header)

    matched = sum(1 for i in col_index if i is not None)
    if matched == 0:
        warnings.append("No blueprint headers matched the selected output spec.")
    elif missing_headers:
        warnings.append("Blueprint is missing output columns: " + ", ".join(missing_headers))

    data_rows = all_rows[header_idx + 1:]
    rows: list[list[str]] = []
    for raw in data_rows:
        cells = [
            _cell_text(raw[src] if (src is not None and src < len(raw)) else None)
            for src in col_index
        ]
        if any(cells):
            rows.append(cells)

    totals = {
        "amount": _sum_column(data_rows, header_row, {"ext_amount", "amount"}),
        "tax": _sum_column(data_rows, header_row, {"ext_tax", "tax", "tax amount", "vat"}),
    }
    return {
        "filename": workbook_path.name,
        "sheet_name": sheet_name,
        "raw_headers": raw_headers,
        "aligned_headers": [c.xlsx_header for c in columns],
        "aligned_rows": rows,
        "row_count": len(rows),
        "sample_rows": rows[:sample_rows],
        "totals": {k: v for k, v in totals.items() if v is not None},
        "warnings": warnings,
        "contract_compatible": not missing_headers and matched == len(spec_headers),
        "missing_output_headers": missing_headers,
        "unexpected_headers": [
            raw_headers[index]
            for index, header in enumerate(header_row)
            if header and header not in spec_headers
        ],
    }


def load_expected_rows(xlsx_path: str | Path, spec: OutputSpec | None) -> list[list[str]]:
    """Backward-compatible wrapper returning rows aligned to ``spec``."""
    return list(load_blueprint_context(xlsx_path, spec)["aligned_rows"])
