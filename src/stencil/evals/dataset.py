"""Eval dataset — labeled cases from the golden corpus.

Each case is one invoice PDF in a corpus layout folder with a committed
``expected/{stem}.xlsx`` answer key and a ``profile.json``. This module lives in
``src`` (not ``tests``) so the runtime eval engine can read cases without importing
the test package; ``tests/corpus_utils.py`` delegates here for the XLSX reader.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from stencil.config import settings


def eval_corpus_dir() -> Path:
    """Root of the labeled eval corpus (defaults to the committed golden corpus)."""
    return Path(settings.eval_corpus_dir)


def read_expected_xlsx_rows(xlsx_path: Path) -> list[dict]:
    """Read an expected-output workbook into a list of {header: value} row dicts."""
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [
        {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
        for row in rows[1:]
        if any(cell is not None and str(cell).strip() for cell in row)
    ]


@dataclass
class EvalCase:
    """One labeled invoice: a PDF + its profile + expected deliverable."""

    case_id: str  # "<layout>/<pdf_stem>"
    layout_id: str
    pdf_path: Path
    profile_path: Path
    expected_xlsx_path: Path

    @property
    def has_expected(self) -> bool:
        return self.expected_xlsx_path.exists()

    def load_profile(self):
        import json

        from stencil.profiles.migrate import migrate_profile_dict
        from stencil.profiles.schema import SupplierProfile

        raw = json.loads(self.profile_path.read_text(encoding="utf-8"))
        return SupplierProfile.model_validate(migrate_profile_dict(raw))

    def expected_rows(self) -> list[dict]:
        return read_expected_xlsx_rows(self.expected_xlsx_path) if self.has_expected else []


def discover_cases(corpus_dir: Path | None = None) -> list[EvalCase]:
    """Every (layout, pdf) with a profile.json — one eval case each, sorted by id."""
    root = corpus_dir or eval_corpus_dir()
    cases: list[EvalCase] = []
    if not root.is_dir():
        return cases
    for folder in sorted(root.iterdir()):
        profile_path = folder / "profile.json"
        invoices_dir = folder / "invoices"
        if not profile_path.exists() or not invoices_dir.is_dir():
            continue
        for pdf in sorted(invoices_dir.glob("*.pdf")):
            cases.append(
                EvalCase(
                    case_id=f"{folder.name}/{pdf.stem}",
                    layout_id=folder.name,
                    pdf_path=pdf,
                    profile_path=profile_path,
                    expected_xlsx_path=folder / "expected" / f"{pdf.stem}.xlsx",
                )
            )
    return cases


def get_case(case_id: str, corpus_dir: Path | None = None) -> EvalCase | None:
    for case in discover_cases(corpus_dir):
        if case.case_id == case_id:
            return case
    return None
